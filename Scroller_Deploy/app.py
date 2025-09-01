import streamlit as st
import pandas as pd
from datetime import datetime
import time
import logging
import re
import json
import requests
from bs4 import BeautifulSoup
from api import SubventionSearcher, FRENCH_REGIONS, EUROPEAN_REGIONS, COLOMBIAN_REGIONS
from recherche_arts_visuels import RechercheArtsVisuels
from european_visual_arts_api import EuropeanVisualArtsFunding
from colombian_visual_arts_api import ColombianVisualArtsFunding
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import os
import io
import os

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create directories for organized file storage
def ensure_directories():
    """Create necessary directories for organized file storage"""
    directories = ['exports/excel', 'exports/logs', 'exports/analysis', 'exports/fixed_logs']
    for directory in directories:
        os.makedirs(directory, exist_ok=True)

# Initialize directories
ensure_directories()

def validate_url(url: str, timeout: int = 10) -> dict:
    """Validate a single URL and return detailed status information"""
    validation_result = {
        'url': url,
        'status': 'Unknown',
        'status_code': None,
        'response_time': None,
        'error_message': None,
        'is_working': False,
        'contains_funding_keywords': False,
        'page_title': None
    }
    
    if not url or not url.startswith(('http://', 'https://')):
        validation_result.update({
            'status': 'Invalid URL',
            'error_message': 'URL is empty or does not start with http/https'
        })
        return validation_result
    
    try:
        start_time = time.time()
        
        # Set headers to mimic a real browser
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        response = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        response_time = (time.time() - start_time) * 1000  # Convert to milliseconds
        
        validation_result['response_time'] = round(response_time, 2)
        validation_result['status_code'] = response.status_code
        
        if response.status_code == 200:
            validation_result['status'] = 'Working'
            validation_result['is_working'] = True
            
            # Try to parse content for additional insights
            try:
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Get page title
                title_tag = soup.find('title')
                if title_tag:
                    validation_result['page_title'] = title_tag.get_text().strip()
                
                # Check for funding-related keywords
                page_text = soup.get_text().lower()
                funding_keywords = [
                    'grant', 'funding', 'subvention', 'aide', 'financement',
                    'scholarship', 'bourse', 'subsidy', 'support', 'program',
                    'programme', 'application', 'deadline', 'eligibility'
                ]
                
                validation_result['contains_funding_keywords'] = any(keyword in page_text for keyword in funding_keywords)
                
            except Exception as parse_error:
                # Content parsing failed, but URL is still working
                validation_result['error_message'] = f"Content parsing failed: {str(parse_error)}"
                
        elif response.status_code == 404:
            validation_result['status'] = '404 Not Found'
            validation_result['error_message'] = 'Page not found'
        elif response.status_code == 403:
            validation_result['status'] = '403 Forbidden'
            validation_result['error_message'] = 'Access forbidden'
        elif response.status_code == 500:
            validation_result['status'] = '500 Server Error'
            validation_result['error_message'] = 'Internal server error'
        else:
            validation_result['status'] = f'HTTP {response.status_code}'
            validation_result['error_message'] = f'HTTP status code: {response.status_code}'
            
    except requests.exceptions.Timeout:
        validation_result.update({
            'status': 'Timeout',
            'error_message': f'Request timed out after {timeout} seconds'
        })
    except requests.exceptions.ConnectionError:
        validation_result.update({
            'status': 'Connection Error',
            'error_message': 'Could not connect to server'
        })
    except requests.exceptions.RequestException as e:
        validation_result.update({
            'status': 'Request Error',
            'error_message': str(e)
        })
    except Exception as e:
        validation_result.update({
            'status': 'Unknown Error',
            'error_message': str(e)
        })
    
    return validation_result

def attempt_url_fix(original_url: str, title: str = "", source: str = "") -> dict:
    """Attempt to fix a broken URL using various strategies"""
    fix_result = {
        'original_url': original_url,
        'title': title,
        'source': source,
        'fix_attempts': [],
        'final_status': 'Not Fixed',
        'working_url': None,
        'fix_strategy': None,
        'issues_found': []
    }
    
    if not original_url:
        fix_result['issues_found'].append("URL is empty")
        return fix_result
    
    # Common URL fixing strategies
    fix_strategies = [
        # Strategy 1: Add missing protocol
        lambda url: f"https://{url}" if not url.startswith(('http://', 'https://')) else url,
        
        # Strategy 2: Fix common typos
        lambda url: url.replace('htpp://', 'http://').replace('htpps://', 'https://').replace('www.www.', 'www.'),
        
        # Strategy 3: Try HTTPS instead of HTTP
        lambda url: url.replace('http://', 'https://') if url.startswith('http://') else url,
        
        # Strategy 4: Try HTTP instead of HTTPS
        lambda url: url.replace('https://', 'http://') if url.startswith('https://') else url,
        
        # Strategy 5: Remove duplicate slashes
        lambda url: url.replace('///', '//').replace('////', '//'),
        
        # Strategy 6: Remove trailing parameters that might be broken
        lambda url: url.split('?')[0] if '?' in url else url,
        
        # Strategy 7: Try without www
        lambda url: url.replace('www.', '') if 'www.' in url else url,
        
        # Strategy 8: Try with www
        lambda url: url.replace('://', '://www.') if '://www.' not in url and '://' in url else url,
        
        # Strategy 9: Try removing trailing slash
        lambda url: url.rstrip('/') if url.endswith('/') else url,
        
        # Strategy 10: Try adding trailing slash
        lambda url: f"{url}/" if not url.endswith('/') and '?' not in url and '#' not in url else url
    ]
    
    strategy_names = [
        "Add missing protocol",
        "Fix common typos", 
        "Try HTTPS instead of HTTP",
        "Try HTTP instead of HTTPS",
        "Remove duplicate slashes",
        "Remove URL parameters",
        "Try without www",
        "Try with www",
        "Remove trailing slash",
        "Add trailing slash"
    ]
    
    # Test original URL first
    original_validation = validate_url(original_url, timeout=5)
    fix_result['fix_attempts'].append({
        'strategy': 'Original URL',
        'url': original_url,
        'status': original_validation['status'],
        'working': original_validation['is_working']
    })
    
    if original_validation['is_working']:
        fix_result['final_status'] = 'Already Working'
        fix_result['working_url'] = original_url
        return fix_result
    
    # Try each fixing strategy
    for i, (strategy_func, strategy_name) in enumerate(zip(fix_strategies, strategy_names)):
        try:
            fixed_url = strategy_func(original_url)
            
            # Skip if URL didn't change
            if fixed_url == original_url:
                continue
                
            # Skip if we already tested this URL
            if any(attempt['url'] == fixed_url for attempt in fix_result['fix_attempts']):
                continue
            
            # Test the fixed URL
            validation = validate_url(fixed_url, timeout=5)
            
            fix_attempt = {
                'strategy': strategy_name,
                'url': fixed_url,
                'status': validation['status'],
                'working': validation['is_working'],
                'response_time': validation.get('response_time'),
                'page_title': validation.get('page_title')
            }
            
            fix_result['fix_attempts'].append(fix_attempt)
            
            # If we found a working URL, stop trying
            if validation['is_working']:
                fix_result['final_status'] = 'Fixed'
                fix_result['working_url'] = fixed_url
                fix_result['fix_strategy'] = strategy_name
                break
                
        except Exception as e:
            fix_result['fix_attempts'].append({
                'strategy': strategy_name,
                'url': 'Error generating URL',
                'status': 'Error',
                'working': False,
                'error': str(e)
            })
    
    # If no fix worked, try domain-specific strategies
    if fix_result['final_status'] == 'Not Fixed':
        domain_fixes = []
        
        # Extract domain for specific fixes
        try:
            from urllib.parse import urlparse
            parsed = urlparse(original_url)
            domain = parsed.netloc.lower()
            
            # Domain-specific fixing strategies
            if 'ec.europa.eu' in domain:
                # European Commission URLs often change structure
                if '/calls-for-proposals' in original_url:
                    domain_fixes.append('https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/home')
                if '/regional_policy' in original_url:
                    domain_fixes.append('https://ec.europa.eu/regional_policy/funding_en')
                    
            elif 'culture.gouv.fr' in domain:
                # French culture ministry URLs
                domain_fixes.append('https://www.culture.gouv.fr/Aides-demarches')
                domain_fixes.append('https://www.culture.gouv.fr/Thematiques/Soutien-a-la-creation-artistique')
                
            elif 'artscouncil' in domain:
                # Arts council URLs often redirect
                if '.org.uk' in domain:
                    domain_fixes.append('https://www.artscouncil.org.uk/funding')
                    
            elif 'kulturfond' in domain or 'kultuur' in domain:
                # Nordic/Cultural fund URLs
                domain_fixes.append('https://www.nordiskkulturfond.org/en/funding')
                
        except:
            pass
        
        # Test domain-specific fixes
        for fixed_url in domain_fixes:
            if not any(attempt['url'] == fixed_url for attempt in fix_result['fix_attempts']):
                validation = validate_url(fixed_url, timeout=5)
                
                fix_attempt = {
                    'strategy': 'Domain-specific fix',
                    'url': fixed_url,
                    'status': validation['status'],
                    'working': validation['is_working'],
                    'response_time': validation.get('response_time'),
                    'page_title': validation.get('page_title')
                }
                
                fix_result['fix_attempts'].append(fix_attempt)
                
                if validation['is_working']:
                    fix_result['final_status'] = 'Fixed'
                    fix_result['working_url'] = fixed_url
                    fix_result['fix_strategy'] = 'Domain-specific fix'
                    break
    
    # Analyze issues found
    if fix_result['final_status'] == 'Not Fixed':
        # Common issues analysis
        if original_url.count('//') > 1:
            fix_result['issues_found'].append("Multiple slashes in URL")
        if not original_url.startswith(('http://', 'https://')):
            fix_result['issues_found'].append("Missing protocol")
        if 'www.www.' in original_url:
            fix_result['issues_found'].append("Duplicate www prefix")
        if any(attempt['status'] == '404 Not Found' for attempt in fix_result['fix_attempts']):
            fix_result['issues_found'].append("Page no longer exists (404)")
        if any(attempt['status'] == 'Connection Error' for attempt in fix_result['fix_attempts']):
            fix_result['issues_found'].append("Server unreachable")
        if any(attempt['status'] == 'Timeout' for attempt in fix_result['fix_attempts']):
            fix_result['issues_found'].append("Server too slow to respond")
        if any(attempt['status'] == '403 Forbidden' for attempt in fix_result['fix_attempts']):
            fix_result['issues_found'].append("Access forbidden by server")
        if len(fix_result['fix_attempts']) > 5:
            fix_result['issues_found'].append("Multiple fix strategies failed")
    
    return fix_result

def create_url_fix_log(broken_urls: list, filename: str) -> str:
    """Create a comprehensive log of URL fixing attempts"""
    try:
        # Ensure directory exists
        fixed_logs_dir = "exports/fixed_logs"
        os.makedirs(fixed_logs_dir, exist_ok=True)
        
        # Full path for fix log file
        log_path = os.path.join(fixed_logs_dir, filename)
        
        if not broken_urls:
            return None
        
        st.info(f"🔧 Attempting to fix {len(broken_urls)} broken URLs...")
        
        fix_results = []
        fixed_count = 0
        
        # Progress indicator
        progress_placeholder = st.empty()
        
        for i, broken_url in enumerate(broken_urls):
            progress_placeholder.text(f"Fixing URL {i+1}/{len(broken_urls)}: {broken_url.get('title', 'Unknown')[:50]}...")
            
            fix_result = attempt_url_fix(
                broken_url['url'],
                broken_url.get('title', ''),
                broken_url.get('source', '')
            )
            
            fix_results.append(fix_result)
            
            if fix_result['final_status'] == 'Fixed':
                fixed_count += 1
        
        progress_placeholder.empty()
        
        # Create detailed fix log
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        fix_log_data = {
            "metadata": {
                "timestamp": timestamp,
                "total_broken_urls": len(broken_urls),
                "successfully_fixed": fixed_count,
                "fix_success_rate": f"{(fixed_count / len(broken_urls) * 100):.1f}%" if broken_urls else "0%",
                "fix_log_version": "1.0"
            },
            "summary": {
                "fixed_urls": fixed_count,
                "already_working": len([r for r in fix_results if r['final_status'] == 'Already Working']),
                "not_fixed": len([r for r in fix_results if r['final_status'] == 'Not Fixed']),
                "common_issues": []
            },
            "fix_results": fix_results,
            "recommendations": []
        }
        
        # Analyze common issues
        all_issues = []
        for result in fix_results:
            all_issues.extend(result.get('issues_found', []))
        
        # Count issue frequency
        issue_counts = {}
        for issue in all_issues:
            issue_counts[issue] = issue_counts.get(issue, 0) + 1
        
        # Sort by frequency
        common_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)
        fix_log_data["summary"]["common_issues"] = [
            {"issue": issue, "count": count, "percentage": f"{count/len(broken_urls)*100:.1f}%"} 
            for issue, count in common_issues[:5]
        ]
        
        # Generate recommendations
        if fixed_count > 0:
            fix_log_data["recommendations"].append(f"Successfully fixed {fixed_count} URLs - update database with new working URLs")
        
        if len([r for r in fix_results if r['final_status'] == 'Not Fixed']) > 0:
            fix_log_data["recommendations"].append("Consider removing permanently broken URLs or finding alternative sources")
        
        if issue_counts.get("Page no longer exists (404)", 0) > len(broken_urls) * 0.5:
            fix_log_data["recommendations"].append("Many pages no longer exist - website restructuring likely occurred")
        
        if issue_counts.get("Server unreachable", 0) > 0:
            fix_log_data["recommendations"].append("Some servers are unreachable - check if organizations still exist")
        
        # Save to JSON file
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump(fix_log_data, f, indent=2, ensure_ascii=False)
        
        # Also create a human-readable text version
        text_log_path = log_path.replace('.json', '.txt')
        with open(text_log_path, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("URL FIX ATTEMPT REPORT\n")
            f.write("="*80 + "\n")
            f.write(f"Generated: {timestamp}\n")
            f.write(f"Total Broken URLs: {len(broken_urls)}\n")
            f.write(f"Successfully Fixed: {fixed_count} ({(fixed_count / len(broken_urls) * 100):.1f}%)\n\n")
            
            f.write("FIXED URLS\n")
            f.write("-"*40 + "\n")
            for result in fix_results:
                if result['final_status'] == 'Fixed':
                    f.write(f"✅ FIXED: {result['title'][:60]}...\n")
                    f.write(f"   Original: {result['original_url']}\n")
                    f.write(f"   Fixed:    {result['working_url']}\n")
                    f.write(f"   Strategy: {result['fix_strategy']}\n")
                    f.write(f"   Source:   {result['source']}\n")
                    f.write("\n")
            
            f.write("STILL BROKEN URLS\n")
            f.write("-"*40 + "\n")
            for result in fix_results:
                if result['final_status'] == 'Not Fixed':
                    f.write(f"❌ NOT FIXED: {result['title'][:60]}...\n")
                    f.write(f"   URL: {result['original_url']}\n")
                    f.write(f"   Issues: {', '.join(result['issues_found'])}\n")
                    f.write(f"   Attempts: {len(result['fix_attempts'])} strategies tried\n")
                    f.write(f"   Source: {result['source']}\n")
                    f.write("\n")
            
            f.write("COMMON ISSUES ANALYSIS\n")
            f.write("-"*40 + "\n")
            for issue_data in fix_log_data["summary"]["common_issues"]:
                f.write(f"• {issue_data['issue']}: {issue_data['count']} URLs ({issue_data['percentage']})\n")
            f.write("\n")
            
            f.write("RECOMMENDATIONS\n")
            f.write("-"*40 + "\n")
            for rec in fix_log_data["recommendations"]:
                f.write(f"• {rec}\n")
        
        # Show results to user
        if fixed_count > 0:
            st.success(f"🔧 Successfully fixed {fixed_count}/{len(broken_urls)} URLs ({(fixed_count / len(broken_urls) * 100):.1f}%)")
        else:
            st.warning("⚠️ No URLs could be automatically fixed")
        
        logger.info(f"URL fix log saved: {log_path}")
        return log_path
        
    except Exception as e:
        logger.error(f"Error creating URL fix log: {e}")
        st.error(f"❌ Error creating URL fix log: {e}")
        return None

# Set page configuration
st.set_page_config(
    page_title="Subvention Research Bot",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    /* Dark theme for main app */
    .stApp {
        background-color: #0e1117;
    }
    
    .main > div {
        background-color: #0e1117;
    }
    
    .stMainBlockContainer {
        background-color: #0e1117;
        padding-top: 1rem;
    }
    
    .main-header {
        background: linear-gradient(90deg, #1f4e79, #2e7bb8);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    
    .search-container {
        background: #262730;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 4px solid #2e7bb8;
        margin-bottom: 2rem;
        color: #fafafa;
    }
    
    .result-card {
        background: #262730;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #28a745;
        margin-bottom: 1rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
        color: #fafafa;
    }
    
    .french-result {
        border-left-color: #007bff;
        background: #1a1d29;
    }
    
    .european-result {
        border-left-color: #ffc107;
        background: #2a251a;
    }
    
    .colombian-result {
        border-left-color: #ff6b35;
        background: #2a1a1a;
    }
    
    .stats-container {
        display: flex;
        justify-content: space-around;
        margin: 1rem 0;
    }
    
    .stat-box {
        background: #262730;
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
        min-width: 150px;
        color: #fafafa;
    }
    
    .stButton > button {
        width: 100%;
        background: linear-gradient(90deg, #28a745, #20c997);
        color: white;
        border: none;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        font-weight: bold;
    }
    
    .stButton > button:hover {
        background: linear-gradient(90deg, #218838, #1ea085);
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    
    /* Dark theme for expanders and other elements */
    .streamlit-expanderHeader {
        background-color: #262730 !important;
        color: #fafafa !important;
    }
    
    .streamlit-expanderContent {
        background-color: #1a1d29 !important;
        color: #fafafa !important;
    }
    
    /* Dark theme for tabs */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #262730;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #262730;
        color: #fafafa;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #2e7bb8 !important;
        color: white !important;
    }
    
    /* Welcome container styling */
    .welcome-container {
        background: #262730;
        padding: 2rem;
        border-radius: 10px;
        border-left: 4px solid #2e7bb8;
        margin-bottom: 2rem;
        color: #fafafa;
    }
    
    .welcome-container h3 {
        color: #2e7bb8;
        margin-bottom: 1rem;
    }
    
    .welcome-container ul {
        color: #fafafa;
    }
    
    .welcome-container li {
        margin-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

def init_session_state():
    """Initialize session state variables"""
    if 'search_results' not in st.session_state:
        st.session_state.search_results = {'french': [], 'european': [], 'colombian': []}
    
    if 'last_search_params' not in st.session_state:
        st.session_state.last_search_params = {}
    
    if 'search_performed' not in st.session_state:
        st.session_state.search_performed = False

def create_styled_excel(results_df: pd.DataFrame, filename: str) -> str:
    """Create a beautifully formatted Excel file"""
    try:
        # Ensure directory exists
        excel_dir = "exports/excel"
        os.makedirs(excel_dir, exist_ok=True)
        
        # Full path for Excel file
        excel_path = os.path.join(excel_dir, filename)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subventions Results"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7BB8', end_color='2E7BB8', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title row
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"Subvention Research Results - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='2E7BB8')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers in row 3 - Enhanced for all result types
        headers = [
            'Title', 'Description', 'Organization/Source', 'Category', 'Type', 
            'Location/Region', 'Funding Amount', 'Amount Min', 'Amount Max',
            'Deadline', 'Days Until Deadline', 'Contact Email', 'Contact Phone', 
            'Website/URL', 'Eligibility', 'Aid Types', 'Targeted Audiences', 
            'Financers', 'Project Manager', 'Additional Info', 'Date Created'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows - Enhanced to handle all result types (French, European, Colombian)
        for row_idx, (_, row) in enumerate(results_df.iterrows(), 4):
            # Handle different data structures for different sources
            organization = row.get('organization', row.get('organisme', row.get('source', '')))
            funding_amount = row.get('funding_amount', row.get('montant', row.get('amount', '')))
            contact_email = row.get('contact', row.get('contact_email', ''))
            website = row.get('website', row.get('url', ''))
            location = row.get('location', row.get('perimeter', row.get('region', '')))
            project_type = row.get('type', row.get('type_projet', ''))
            eligibility = row.get('eligibility', row.get('eligibilite', ''))
            
            for col_idx, value in enumerate([
                row.get('title', ''),
                row.get('description', ''),
                organization,
                row.get('category', ''),
                project_type,
                location,
                funding_amount,
                row.get('amount_min', ''),
                row.get('amount_max', ''),
                row.get('deadline', ''),
                row.get('days_until_deadline', ''),
                contact_email,
                row.get('contact_tel', row.get('contact_phone', '')),
                website,
                eligibility,
                row.get('aid_types', ''),
                row.get('targeted_audiences', ''),
                row.get('financers', ''),
                row.get('project_manager', ''),
                row.get('additional_info', row.get('criteres', '')),
                row.get('date_created', '')
            ], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = str(value) if value else ''
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
                
                # Add category-specific coloring
                if col_idx == 3:  # Category column
                    if 'French' in str(value):
                        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                    elif 'European' in str(value):
                        cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        
        # Adjust column widths for enhanced headers
        column_widths = [40, 60, 25, 20, 20, 25, 20, 15, 15, 15, 10, 30, 15, 40, 35, 25, 25, 25, 30, 35, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Set row heights
        for row in range(3, ws.max_row + 1):
            ws.row_dimensions[row].height = 60
        
        # Save file
        wb.save(excel_path)
        logger.info(f"Excel file saved: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return None

def save_filtered_excel(filter_type: str = "working_urls"):
    """Save filtered Excel file with only specific types of results"""
    try:
        if 'search_results' not in st.session_state or not st.session_state.search_results:
            st.error("❌ No search results to filter. Please perform a search first.")
            return None
        
        results = st.session_state.search_results
        all_results = results.get('french', []) + results.get('european', []) + results.get('colombian', [])
        
        if not all_results:
            st.error("❌ No results to filter.")
            return None
        
        # Filter results based on type
        if filter_type == "working_urls":
            st.info("🔍 Validating URLs to filter working projects...")
            
            # Validate URLs for filtering
            filtered_results = []
            progress_placeholder = st.empty()
            
            for i, result in enumerate(all_results):
                progress_placeholder.text(f"Validating URL {i+1}/{len(all_results)}: {result.get('title', 'Unknown')[:50]}...")
                
                url = result.get('url', '')
                if url:
                    validation = validate_url(url, timeout=5)
                    if validation['is_working']:
                        # Add validation info to result
                        result_copy = result.copy()
                        result_copy['url_status'] = validation['status']
                        result_copy['url_response_time'] = validation.get('response_time', 0)
                        result_copy['url_page_title'] = validation.get('page_title', '')
                        result_copy['url_contains_funding'] = validation.get('contains_funding_keywords', False)
                        filtered_results.append(result_copy)
                else:
                    # If no URL, skip this result
                    continue
            
            progress_placeholder.empty()
            
            if not filtered_results:
                st.warning("⚠️ No projects with working URLs found.")
                return None
            
            st.success(f"✅ Found {len(filtered_results)} projects with working URLs out of {len(all_results)} total projects")
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"working_urls_only_{timestamp}.xlsx"
            
        else:
            st.error("❌ Unknown filter type")
            return None
        
        # Convert to DataFrame
        df = pd.DataFrame(filtered_results)
        
        # Create the Excel file
        excel_path = create_filtered_excel(df, filename, filter_type)
        
        if excel_path:
            # Provide download link
            with open(excel_path, "rb") as file:
                st.download_button(
                    label="📥 Download Filtered Excel File",
                    data=file.read(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{filter_type}_{timestamp}"
                )
            
            st.success(f"✅ Filtered Excel file created: {filename}")
            return excel_path
        
    except Exception as e:
        logger.error(f"Error creating filtered Excel: {e}")
        st.error(f"❌ Error creating filtered Excel: {e}")
        return None

def create_filtered_excel(results_df: pd.DataFrame, filename: str, filter_type: str) -> str:
    """Create a beautifully formatted Excel file for filtered results"""
    try:
        # Ensure directory exists
        excel_dir = "exports/excel"
        os.makedirs(excel_dir, exist_ok=True)
        
        # Full path for Excel file
        excel_path = os.path.join(excel_dir, filename)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if filter_type == "working_urls":
            ws.title = "Working URLs Only"
            title_text = "Projects with Verified Working URLs"
            header_color = '28A745'  # Green for working URLs
        else:
            ws.title = "Filtered Results"
            title_text = "Filtered Results"
            header_color = '2E7BB8'
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title row
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f"{title_text} - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(name='Arial', size=16, bold=True, color=header_color)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add filter info row
        ws.merge_cells('A2:P2')
        info_cell = ws['A2']
        if filter_type == "working_urls":
            info_cell.value = f"✅ This file contains {len(results_df)} projects with verified working URLs"
        info_cell.font = Font(name='Arial', size=12, color='666666')
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Enhanced headers for all result types
        headers = [
            'Title', 'Description', 'Organization/Source', 'Category', 'Type', 
            'Location/Region', 'Funding Amount', 'Amount Min', 'Amount Max',
            'Deadline', 'Days Until Deadline', 'Contact Email', 'Contact Phone', 
            'Website/URL', 'Eligibility', 'Aid Types', 'Targeted Audiences', 
            'Financers', 'Project Manager', 'Additional Info', 'Date Created'
        ]
        
        # Add URL validation headers if working URLs filter
        if filter_type == "working_urls":
            headers.extend(['URL Status', 'Response Time (ms)', 'Page Title', 'Contains Funding Info'])
        
        # Add headers in row 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows - Enhanced to handle all result types
        for row_idx, (_, row) in enumerate(results_df.iterrows(), 5):
            # Handle different data structures for different sources
            organization = row.get('organization', row.get('organisme', row.get('source', '')))
            funding_amount = row.get('funding_amount', row.get('montant', row.get('amount', '')))
            contact_email = row.get('contact', row.get('contact_email', ''))
            website = row.get('website', row.get('url', ''))
            location = row.get('location', row.get('perimeter', row.get('region', '')))
            project_type = row.get('type', row.get('type_projet', ''))
            eligibility = row.get('eligibility', row.get('eligibilite', ''))
            
            data_values = [
                row.get('title', ''),
                row.get('description', ''),
                organization,
                row.get('category', ''),
                project_type,
                location,
                funding_amount,
                row.get('amount_min', ''),
                row.get('amount_max', ''),
                row.get('deadline', ''),
                row.get('days_until_deadline', ''),
                contact_email,
                row.get('contact_tel', row.get('contact_phone', '')),
                website,
                eligibility,
                row.get('aid_types', ''),
                row.get('targeted_audiences', ''),
                row.get('financers', ''),
                row.get('project_manager', ''),
                row.get('additional_info', row.get('criteres', '')),
                row.get('date_created', '')
            ]
            
            # Add URL validation data if working URLs filter
            if filter_type == "working_urls":
                data_values.extend([
                    row.get('url_status', 'Working'),
                    row.get('url_response_time', ''),
                    row.get('url_page_title', ''),
                    'Yes' if row.get('url_contains_funding', False) else 'No'
                ])
            
            for col_idx, value in enumerate(data_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = str(value) if value else ''
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
                
                # Add category-specific coloring
                if col_idx == 3:  # Category column
                    if 'French' in str(value):
                        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                    elif 'European' in str(value):
                        cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                
                # Highlight working URL status
                if filter_type == "working_urls" and col_idx == len(data_values) - 3:  # URL Status column
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        
        # Adjust column widths for enhanced headers
        base_widths = [40, 60, 25, 20, 20, 25, 20, 15, 15, 15, 10, 30, 15, 40, 35, 25, 25, 25, 30, 35, 12]
        if filter_type == "working_urls":
            base_widths.extend([15, 12, 40, 15])  # URL validation columns
        
        for col, width in enumerate(base_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Set row heights
        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 60
        
        # Save file
        wb.save(excel_path)
        logger.info(f"Filtered Excel file saved: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"Error creating filtered Excel file: {e}")
        return None

def create_analysis_log(results: dict, search_params: dict, filename: str, auto_fix_urls: bool = False) -> str:
    """Create comprehensive analysis log with statistics, insights, and URL validation"""
    try:
        # Ensure directory exists
        logs_dir = "exports/logs"
        os.makedirs(logs_dir, exist_ok=True)
        
        # Full path for log file
        log_path = os.path.join(logs_dir, filename)
        
        french_results = results.get('french', [])
        european_results = results.get('european', [])
        colombian_results = results.get('colombian', [])
        all_results = french_results + european_results + colombian_results
        
        # Calculate comprehensive statistics
        total_results = len(all_results)
        total_french = len(french_results)
        total_european = len(european_results)
        total_colombian = len(colombian_results)
        
        # Urgency analysis
        urgent_7 = [r for r in all_results if r.get('days_until_deadline', 999) <= 7]
        urgent_30 = [r for r in all_results if r.get('days_until_deadline', 999) <= 30]
        medium_90 = [r for r in all_results if 30 < r.get('days_until_deadline', 999) <= 90]
        distant = [r for r in all_results if r.get('days_until_deadline', 999) > 90]
        
        # Financial analysis
        amounts = []
        for result in all_results:
            try:
                amount_max = result.get('amount_max', '0')
                if isinstance(amount_max, str):
                    amount_max = amount_max.replace(',', '').replace('€', '').replace('EUR', '').strip()
                if amount_max and amount_max.isdigit():
                    amounts.append(float(amount_max))
            except:
                continue
        
        total_funding = sum(amounts) if amounts else 0
        avg_funding = total_funding / len(amounts) if amounts else 0
        max_funding = max(amounts) if amounts else 0
        min_funding = min(amounts) if amounts else 0
        
        # Geographic analysis
        countries = {}
        sources = {}
        for result in all_results:
            # Count by country/region
            country = result.get('perimeter', 'Unknown')
            if country in countries:
                countries[country] += 1
            else:
                countries[country] = 1
            
            # Count by source
            source = result.get('source', 'Unknown')
            if source in sources:
                sources[source] += 1
            else:
                sources[source] = 1
        
        # URL VALIDATION ANALYSIS
        st.info("🔍 Validating URLs... This may take a moment.")
        url_validation_results = []
        working_urls = []
        broken_urls = []
        
        urls_to_validate = []
        for result in all_results:
            url = result.get('url', '')
            if url and url not in [r['url'] for r in urls_to_validate]:
                urls_to_validate.append({
                    'url': url,
                    'title': result.get('title', 'Unknown'),
                    'source': result.get('source', 'Unknown'),
                    'country': result.get('perimeter', 'Unknown')
                })
        
        # Validate URLs with progress indication
        total_urls = len(urls_to_validate)
        if total_urls > 0:
            progress_placeholder = st.empty()
            
            for i, url_info in enumerate(urls_to_validate):
                progress_placeholder.text(f"Validating URL {i+1}/{total_urls}: {url_info['title'][:50]}...")
                
                validation = validate_url(url_info['url'])
                validation.update({
                    'title': url_info['title'],
                    'source': url_info['source'],
                    'country': url_info['country']
                })
                
                url_validation_results.append(validation)
                
                if validation['is_working']:
                    working_urls.append(validation)
                else:
                    broken_urls.append(validation)
            
            progress_placeholder.empty()
        
        # URL statistics
        total_urls_tested = len(url_validation_results)
        working_count = len(working_urls)
        broken_count = len(broken_urls)
        working_percentage = (working_count / total_urls_tested * 100) if total_urls_tested > 0 else 0
        
        # Group broken URLs by error type
        error_breakdown = {}
        for broken_url in broken_urls:
            error_type = broken_url['status']
            if error_type in error_breakdown:
                error_breakdown[error_type] += 1
            else:
                error_breakdown[error_type] = 1
        
        # Create detailed analysis log
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        analysis_data = {
            "metadata": {
                "timestamp": timestamp,
                "search_parameters": search_params,
                "analysis_version": "2.0_with_url_validation"
            },
            "executive_summary": {
                "total_opportunities": total_results,
                "french_opportunities": total_french,
                "european_opportunities": total_european,
                "colombian_opportunities": total_colombian,
                "total_funding_available": f"€{total_funding:,.2f}",
                "average_funding": f"€{avg_funding:,.2f}",
                "highest_funding": f"€{max_funding:,.2f}",
                "lowest_funding": f"€{min_funding:,.2f}"
            },
            "urgency_analysis": {
                "very_urgent_7_days": {
                    "count": len(urgent_7),
                    "percentage": f"{(len(urgent_7)/total_results*100):.1f}%" if total_results > 0 else "0%",
                    "opportunities": [{"title": r.get('title', ''), "deadline": r.get('deadline', ''), "amount": r.get('amount', ''), "url": r.get('url', '')} for r in urgent_7]
                },
                "urgent_30_days": {
                    "count": len(urgent_30),
                    "percentage": f"{(len(urgent_30)/total_results*100):.1f}%" if total_results > 0 else "0%",
                    "opportunities": [{"title": r.get('title', ''), "deadline": r.get('deadline', ''), "amount": r.get('amount', ''), "url": r.get('url', '')} for r in urgent_30]
                },
                "medium_90_days": {
                    "count": len(medium_90),
                    "percentage": f"{(len(medium_90)/total_results*100):.1f}%" if total_results > 0 else "0%"
                },
                "distant_future": {
                    "count": len(distant),
                    "percentage": f"{(len(distant)/total_results*100):.1f}%" if total_results > 0 else "0%"
                }
            },
            "url_validation_analysis": {
                "summary": {
                    "total_urls_tested": total_urls_tested,
                    "working_urls": working_count,
                    "broken_urls": broken_count,
                    "success_rate": f"{working_percentage:.1f}%",
                    "validation_timestamp": timestamp
                },
                "working_urls": [
                    {
                        "url": url["url"],
                        "title": url["title"],
                        "source": url["source"],
                        "country": url["country"],
                        "status": url["status"],
                        "response_time_ms": url["response_time"],
                        "page_title": url["page_title"],
                        "contains_funding_info": url["contains_funding_keywords"]
                    } for url in working_urls
                ],
                "broken_urls": [
                    {
                        "url": url["url"],
                        "title": url["title"],
                        "source": url["source"],
                        "country": url["country"],
                        "status": url["status"],
                        "error_message": url["error_message"],
                        "response_time_ms": url["response_time"]
                    } for url in broken_urls
                ],
                "error_breakdown": error_breakdown,
                "quality_insights": {
                    "fastest_response": min([u['response_time'] for u in working_urls if u['response_time']], default=0),
                    "slowest_response": max([u['response_time'] for u in working_urls if u['response_time']], default=0),
                    "avg_response_time": sum([u['response_time'] for u in working_urls if u['response_time']]) / len([u for u in working_urls if u['response_time']]) if working_urls else 0,
                    "funding_content_rate": f"{len([u for u in working_urls if u['contains_funding_keywords']]) / len(working_urls) * 100:.1f}%" if working_urls else "0%"
                }
            },
            "geographic_distribution": {
                "by_country_region": dict(sorted(countries.items(), key=lambda x: x[1], reverse=True)),
                "top_regions": list(sorted(countries.items(), key=lambda x: x[1], reverse=True))[:5]
            },
            "source_analysis": {
                "by_source": dict(sorted(sources.items(), key=lambda x: x[1], reverse=True)),
                "source_diversity": len(sources)
            },
            "funding_analysis": {
                "total_opportunities_with_amounts": len(amounts),
                "funding_ranges": {
                    "under_10k": len([a for a in amounts if a < 10000]),
                    "10k_to_50k": len([a for a in amounts if 10000 <= a < 50000]),
                    "50k_to_100k": len([a for a in amounts if 50000 <= a < 100000]),
                    "over_100k": len([a for a in amounts if a >= 100000])
                }
            },
            "recommendations": {
                "priority_actions": [],
                "strategy_insights": [],
                "follow_up_needed": [],
                "url_maintenance": []
            }
        }
        
        # Add intelligent recommendations
        if len(urgent_7) > 0:
            analysis_data["recommendations"]["priority_actions"].append(f"IMMEDIATE ACTION: {len(urgent_7)} opportunities close within 7 days - prioritize applications")
        
        if len(urgent_30) > 0:
            analysis_data["recommendations"]["priority_actions"].append(f"HIGH PRIORITY: {len(urgent_30)} opportunities close within 30 days - prepare applications soon")
        
        if avg_funding > 50000:
            analysis_data["recommendations"]["strategy_insights"].append(f"High-value opportunities available (avg €{avg_funding:,.0f}) - focus on quality applications")
        
        if total_european > total_french:
            analysis_data["recommendations"]["strategy_insights"].append("European opportunities dominate - consider cross-border collaboration")
        elif total_french > total_european:
            analysis_data["recommendations"]["strategy_insights"].append("French opportunities dominate - leverage local networks")
        
        # Top funding regions
        if countries:
            top_region = max(countries.items(), key=lambda x: x[1])
            analysis_data["recommendations"]["strategy_insights"].append(f"Most opportunities in {top_region[0]} ({top_region[1]} grants) - consider regional focus")
        
        # URL-specific recommendations
        if working_percentage < 70:
            analysis_data["recommendations"]["url_maintenance"].append(f"URL success rate is {working_percentage:.1f}% - needs improvement")
        
        if broken_count > 0:
            analysis_data["recommendations"]["url_maintenance"].append(f"{broken_count} broken URLs found - update database")
            
            # Most common error
            if error_breakdown:
                most_common_error = max(error_breakdown.items(), key=lambda x: x[1])
                analysis_data["recommendations"]["url_maintenance"].append(f"Most common error: {most_common_error[0]} ({most_common_error[1]} URLs)")
        
        if working_count > 0:
            analysis_data["recommendations"]["follow_up_needed"].append(f"{working_count} verified working URLs - prioritize these opportunities")
        
        # Save to JSON file
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump(analysis_data, f, indent=2, ensure_ascii=False)
        
        # Also create a human-readable text version with URL details
        text_log_path = log_path.replace('.json', '.txt')
        with open(text_log_path, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("SUBVENTION SEARCH ANALYSIS REPORT\n")
            f.write("="*80 + "\n")
            f.write(f"Generated: {timestamp}\n")
            f.write(f"Search Type: {search_params.get('search_type', 'General')}\n")
            f.write(f"Keywords: {search_params.get('keywords', 'N/A')}\n")
            f.write(f"Region: {search_params.get('region', 'All')}\n\n")
            
            f.write("EXECUTIVE SUMMARY\n")
            f.write("-"*40 + "\n")
            f.write(f"Total Opportunities Found: {total_results}\n")
            f.write(f"  • French: {total_french}\n")
            f.write(f"  • European: {total_european}\n")
            f.write(f"  • Colombian: {total_colombian}\n")
            f.write(f"Total Funding Available: €{total_funding:,.2f}\n")
            f.write(f"Average Funding: €{avg_funding:,.2f}\n")
            f.write(f"Funding Range: €{min_funding:,.0f} - €{max_funding:,.0f}\n\n")
            
            f.write("URL VALIDATION RESULTS\n")
            f.write("-"*40 + "\n")
            f.write(f"Total URLs Tested: {total_urls_tested}\n")
            f.write(f"Working URLs: {working_count} ({working_percentage:.1f}%)\n")
            f.write(f"Broken URLs: {broken_count} ({100-working_percentage:.1f}%)\n\n")
            
            if working_urls:
                f.write("✅ WORKING URLS:\n")
                for url in working_urls:
                    f.write(f"  • {url['title'][:60]}...\n")
                    f.write(f"    URL: {url['url']}\n")
                    f.write(f"    Status: {url['status']} ({url['response_time']:.0f}ms)\n")
                    f.write(f"    Source: {url['source']} | Country: {url['country']}\n")
                    if url['page_title']:
                        f.write(f"    Page: {url['page_title'][:100]}\n")
                    f.write(f"    Funding Content: {'Yes' if url['contains_funding_keywords'] else 'No'}\n")
                    f.write("\n")
            
            if broken_urls:
                f.write("❌ BROKEN URLS:\n")
                for url in broken_urls:
                    f.write(f"  • {url['title'][:60]}...\n")
                    f.write(f"    URL: {url['url']}\n")
                    f.write(f"    Status: {url['status']}\n")
                    f.write(f"    Error: {url['error_message']}\n")
                    f.write(f"    Source: {url['source']} | Country: {url['country']}\n")
                    f.write("\n")
            
            f.write("URGENCY BREAKDOWN\n")
            f.write("-"*40 + "\n")
            f.write(f"🔴 Very Urgent (≤7 days): {len(urgent_7)} ({(len(urgent_7)/total_results*100):.1f}%)\n")
            f.write(f"🟠 Urgent (8-30 days): {len(urgent_30)} ({(len(urgent_30)/total_results*100):.1f}%)\n")
            f.write(f"🟡 Medium (31-90 days): {len(medium_90)} ({(len(medium_90)/total_results*100):.1f}%)\n")
            f.write(f"🟢 Distant (>90 days): {len(distant)} ({(len(distant)/total_results*100):.1f}%)\n\n")
            
            if urgent_7:
                f.write("IMMEDIATE ACTION REQUIRED (≤7 DAYS):\n")
                for opp in urgent_7:
                    f.write(f"  • {opp.get('title', 'N/A')} - Deadline: {opp.get('deadline', 'N/A')} - Amount: {opp.get('amount', 'N/A')}\n")
                    f.write(f"    URL: {opp.get('url', 'N/A')}\n")
                f.write("\n")
            
            f.write("GEOGRAPHIC DISTRIBUTION\n")
            f.write("-"*40 + "\n")
            for region, count in sorted(countries.items(), key=lambda x: x[1], reverse=True):
                percentage = (count/total_results*100) if total_results > 0 else 0
                f.write(f"{region}: {count} opportunities ({percentage:.1f}%)\n")
            f.write("\n")
            
            f.write("FUNDING ANALYSIS\n")
            f.write("-"*40 + "\n")
            under_10k = len([a for a in amounts if a < 10000])
            f10k_50k = len([a for a in amounts if 10000 <= a < 50000])
            f50k_100k = len([a for a in amounts if 50000 <= a < 100000])
            over_100k = len([a for a in amounts if a >= 100000])
            
            f.write(f"Under €10,000: {under_10k} opportunities\n")
            f.write(f"€10,000 - €50,000: {f10k_50k} opportunities\n")
            f.write(f"€50,000 - €100,000: {f50k_100k} opportunities\n")
            f.write(f"Over €100,000: {over_100k} opportunities\n\n")
            
            f.write("RECOMMENDATIONS\n")
            f.write("-"*40 + "\n")
            for rec in analysis_data["recommendations"]["priority_actions"]:
                f.write(f"• {rec}\n")
            for rec in analysis_data["recommendations"]["strategy_insights"]:
                f.write(f"• {rec}\n")
            for rec in analysis_data["recommendations"]["follow_up_needed"]:
                f.write(f"• {rec}\n")
            for rec in analysis_data["recommendations"]["url_maintenance"]:
                f.write(f"• {rec}\n")
        
        # Show URL validation summary to user
        if total_urls_tested > 0:
            st.success(f"✅ URL Validation Complete: {working_count}/{total_urls_tested} URLs working ({working_percentage:.1f}%)")
            
            if broken_count > 0:
                st.warning(f"⚠️ Found {broken_count} broken URLs - details saved in analysis log")
                
                # Auto-fix broken URLs if enabled
                if auto_fix_urls and broken_urls:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    fix_log_filename = f"url_fixes_{timestamp}.json"
                    fix_log_path = create_url_fix_log(broken_urls, fix_log_filename)
                    
                    if fix_log_path:
                        st.info(f"🔧 URL fixing results saved to: exports/fixed_logs/{fix_log_filename}")
        
        logger.info(f"Analysis log with URL validation saved: {log_path}")
        return log_path
        
    except Exception as e:
        logger.error(f"Error creating analysis log: {e}")
        st.error(f"❌ Error creating analysis log: {e}")
        return None

def display_search_results(results: dict):
    """Display search results with enhanced dashboard and formatting"""
    french_results = results.get('french', [])
    european_results = results.get('european', [])
    colombian_results = results.get('colombian', [])
    total_results = len(french_results) + len(european_results) + len(colombian_results)
    
    if total_results == 0:
        st.warning("⚠️ No results found. Try different keywords or broader search criteria.")
        return
    
    # Enhanced dashboard with urgency and amount analysis
    st.markdown("## 📊 **Tableau de Bord des Subventions**")
    
    # Calculate statistics
    all_results = french_results + european_results + colombian_results
    urgent_count = len([r for r in all_results if r.get('days_until_deadline', 999) <= 30])
    total_amount = sum([float(r.get('amount_max', '0').replace(',', '')) for r in all_results if r.get('amount_max', '').replace(',', '').isdigit()])
    avg_amount = total_amount / len(all_results) if all_results else 0
    
    # Upcoming deadlines (next 7 days)
    very_urgent = [r for r in all_results if r.get('days_until_deadline', 999) <= 7]
    
    # Display enhanced statistics
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    with col1:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #007bff; margin: 0;">{}</h3>
            <p style="margin: 0;">🇫🇷 Subventions<br>Françaises</p>
        </div>
        """.format(len(french_results)), unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #ffc107; margin: 0;">{}</h3>
            <p style="margin: 0;">🇪🇺 Subventions<br>Européennes</p>
        </div>
        """.format(len(european_results)), unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #ff6b35; margin: 0;">{}</h3>
            <p style="margin: 0;">🇨🇴 Financiación<br>Colombiana</p>
        </div>
        """.format(len(colombian_results)), unsafe_allow_html=True)
    
    with col4:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #dc3545; margin: 0;">{}</h3>
            <p style="margin: 0;">🔥 Urgentes<br>(≤30 jours)</p>
        </div>
        """.format(urgent_count), unsafe_allow_html=True)
    
    with col5:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #28a745; margin: 0;">€{:,.0f}</h3>
            <p style="margin: 0;">💰 Montant<br>Moyen</p>
        </div>
        """.format(avg_amount), unsafe_allow_html=True)
    
    with col6:
        st.markdown("""
        <div class="stat-box">
            <h3 style="color: #6f42c1; margin: 0;">{}</h3>
            <p style="margin: 0;">⚡ Très Urgentes<br>(≤7 jours)</p>
        </div>
        """.format(len(very_urgent)), unsafe_allow_html=True)
    
    # Urgent deadlines alert
    if very_urgent:
        st.markdown("### 🚨 **Alertes - Dates Limites Imminentes**")
        for urgent_item in very_urgent:
            days_left = urgent_item.get('days_until_deadline', 0)
            st.error(f"🔴 **{urgent_item.get('title', 'N/A')}** - Date limite dans {days_left} jour{'s' if days_left != 1 else ''} ({urgent_item.get('deadline', 'N/A')})")
    
    # Create tabs for different result types
    tabs_config = []
    if french_results:
        tabs_config.append("🇫🇷 Subventions Françaises")
    if european_results:
        tabs_config.append("🇪🇺 Subventions Européennes") 
    if colombian_results:
        tabs_config.append("�� Financiación Colombiana")
    
    # Always add urgency and all results tabs
    tabs_config.extend(["⏰ Par Urgence", "📊 Tous les Résultats"])
    
    # Create tabs dynamically
    tabs = st.tabs(tabs_config)
    tab_index = 0
    
    # Display French results
    if french_results:
        with tabs[tab_index]:
            st.markdown(f"**🇫🇷 {len(french_results)} subventions françaises trouvées:**")
            for i, result in enumerate(french_results, 1):
                display_result_card(result, f"fr_{i}", "french")
        tab_index += 1
    
    # Display European results
    if european_results:
        with tabs[tab_index]:
            st.markdown(f"**🇪🇺 {len(european_results)} subventions européennes trouvées:**")
            for i, result in enumerate(european_results, 1):
                display_result_card(result, f"eu_{i}", "european")
        tab_index += 1
    
    # Display Colombian results
    if colombian_results:
        with tabs[tab_index]:
            st.markdown(f"**🇨🇴 {len(colombian_results)} oportunidades de financiación colombianas encontradas:**")
            for i, result in enumerate(colombian_results, 1):
                display_result_card(result, f"co_{i}", "colombian")
        tab_index += 1
    
    # Display by urgency
    with tabs[tab_index]:
        st.markdown("### ⏰ **Subventions Triées par Urgence**")
        
        # Group by urgency
        urgent_7 = [r for r in all_results if r.get('days_until_deadline', 999) <= 7]
        urgent_30 = [r for r in all_results if 7 < r.get('days_until_deadline', 999) <= 30]
        medium_90 = [r for r in all_results if 30 < r.get('days_until_deadline', 999) <= 90]
        distant = [r for r in all_results if r.get('days_until_deadline', 999) > 90]
        
        global_index = 0
        
        if urgent_7:
            st.markdown("#### 🔴 **Très Urgent (≤7 jours)**")
            for result in urgent_7:
                global_index += 1
                if result in french_results:
                    result_type = "french"
                elif result in european_results:
                    result_type = "european"
                else:
                    result_type = "colombian"
                display_result_card(result, f"urgent_{global_index}", result_type)
        
        if urgent_30:
            st.markdown("#### 🟠 **Urgent (8-30 jours)**")
            for result in urgent_30:
                global_index += 1
                if result in french_results:
                    result_type = "french"
                elif result in european_results:
                    result_type = "european"
                else:
                    result_type = "colombian"
                display_result_card(result, f"urgent_{global_index}", result_type)
        
        if medium_90:
            st.markdown("#### 🟡 **Modéré (31-90 jours)**")
            for result in medium_90:
                global_index += 1
                if result in french_results:
                    result_type = "french"
                elif result in european_results:
                    result_type = "european"
                else:
                    result_type = "colombian"
                display_result_card(result, f"urgent_{global_index}", result_type)
        
        if distant:
            st.markdown("#### 🟢 **Pas urgent (>90 jours)**")
            for result in distant:
                global_index += 1
                if result in french_results:
                    result_type = "french"
                elif result in european_results:
                    result_type = "european"
                else:
                    result_type = "colombian"
                display_result_card(result, f"urgent_{global_index}", result_type)
    
    tab_index += 1
    
    # Display all results
    with tabs[tab_index]:
        st.markdown(f"**📊 Tous les {total_results} résultats:**")
        for i, result in enumerate(all_results, 1):
            if result in french_results:
                result_type = "french"
            elif result in european_results:
                result_type = "european"
            else:
                result_type = "colombian"
            # Use a different prefix for all results tab to avoid conflicts
            display_result_card(result, f"all_{i}", result_type)
    
    # Download section
    st.markdown("---")
    st.markdown("### 📥 Download Results")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("📊 Download Full Excel Report", help="Download complete results including all projects"):
            save_results_to_excel(download=True)
    
    with col2:
        if st.button("✅ Download Working URLs Only", help="Download only projects with verified working URLs"):
            save_filtered_excel(filter_type="working_urls")
    
    with col3:
        if st.button("📋 Generate Analysis Report"):
            save_results_to_excel_and_logs(False, True, False)
    
    # File organization info
    st.markdown("---")
    st.markdown("### 📁 **File Organization**")
    st.info("""
    **Your files are automatically organized in:**
    - 📊 **Excel Reports**: `exports/excel/` - Detailed spreadsheets with all funding data
    - ✅ **Filtered Excel**: `exports/excel/` - Curated reports with only working URLs
    - 📋 **Analysis Logs**: `exports/logs/` - JSON and text analysis with statistics and recommendations
    - 🔧 **URL Fix Logs**: `exports/fixed_logs/` - Detailed URL repair attempts and results
    """)
    
    # Add download tips
    st.markdown("### 💡 **Download Tips**")
    st.success("""
    **📥 Download Options Available:**
    - **Full Excel Report**: Complete dataset with all projects
    - **Working URLs Only**: Filtered dataset with verified functional links
    - **Analysis Report**: Comprehensive statistics and URL validation insights
    
    **🔍 URL Quality Features:**
    - Automatic URL validation during analysis
    - Broken URL detection and categorization
    - Optional URL repair attempts with detailed logs
    """)

def display_result_card(result: dict, index, result_type: str):
    """Display a single result card with enhanced information"""
    if result_type == "french":
        type_class = "french-result"
        flag = "🇫🇷"
    elif result_type == "european":
        type_class = "european-result"
        flag = "��"
    else:  # colombian
        type_class = "colombian-result"
        flag = "��"
    
    # Calculate urgency indicator
    days_until = result.get('days_until_deadline', 999)
    if days_until <= 7:
        urgency_color = "#dc3545"  # Red - Very urgent
        urgency_text = "🔴 URGENT"
    elif days_until <= 30:
        urgency_color = "#fd7e14"  # Orange - Urgent
        urgency_text = "🟠 Urgent"
    elif days_until <= 90:
        urgency_color = "#ffc107"  # Yellow - Medium
        urgency_text = "🟡 Modéré"
    else:
        urgency_color = "#28a745"  # Green - No rush
        urgency_text = "🟢 Pas urgent"
    
    title_with_urgency = f"{flag} {index}. {result.get('title', 'No title')}"
    if days_until < 999:
        title_with_urgency += f" • {urgency_text} ({days_until} jours)"
    
    with st.expander(title_with_urgency, expanded=False):
        # Main information layout
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown(f"**📄 Description:** {result.get('description', 'No description available')}")
            
            # Handle different source field names based on result type
            if result_type == "colombian":
                st.markdown(f"**🏛️ Organisation:** {result.get('organization', 'N/A')}")
                if result.get('category'):
                    st.markdown(f"**📋 Catégorie:** {result.get('category')}")
                if result.get('eligibility'):
                    st.markdown(f"**🎯 Éligibilité:** {result.get('eligibility')}")
                if result.get('location'):
                    st.markdown(f"**📍 Localisation:** {result.get('location')}")
                if result.get('type'):
                    st.markdown(f"**📋 Type:** {result.get('type')}")
            else:
                # French/European results
                st.markdown(f"**🏛️ Source:** {result.get('source', 'N/A')}")
                if result.get('aid_types'):
                    st.markdown(f"**📋 Types d'aide:** {result.get('aid_types')}")
                if result.get('targeted_audiences'):
                    st.markdown(f"**🎯 Public cible:** {result.get('targeted_audiences')}")
                if result.get('financers'):
                    st.markdown(f"**🏦 Financeurs:** {result.get('financers')}")
        
        with col2:
            # Dates and deadline information
            st.markdown("### 📅 **Dates importantes**")
            if result.get('date_created'):
                st.markdown(f"**📅 Publication:** {result.get('date_created')}")
            if result.get('deadline'):
                deadline_color = urgency_color
                st.markdown(f"**⏰ Date limite:** <span style='color: {deadline_color}; font-weight: bold;'>{result.get('deadline')}</span>", unsafe_allow_html=True)
                if days_until < 999:
                    st.markdown(f"**⏳ Jours restants:** <span style='color: {deadline_color}; font-weight: bold;'>{days_until} jours</span>", unsafe_allow_html=True)
            
            # Amount information - handle different field names
            funding_amount = result.get('funding_amount') or result.get('amount')
            if funding_amount:
                st.markdown("### 💰 **Financement**")
                st.markdown(f"**💰 Montant:** {funding_amount}")
                if result.get('amount_min') and result.get('amount_max'):
                    st.markdown(f"**📊 Fourchette:** €{result.get('amount_min')} - €{result.get('amount_max')}")
            
            # Region information
            if result.get('perimeter'):
                st.markdown(f"**🌍 Périmètre:** {result.get('perimeter')}")
        
        # Contact and project manager information
        has_contact_info = (result.get('project_manager') or result.get('contact_info') or 
                           result.get('contact_email') or result.get('contact_tel') or 
                           result.get('contact_adresse') or result.get('contact'))
        
        if has_contact_info:
            st.markdown("---")
            st.markdown("### 👤 **Contact et informations**")
            
            contact_col1, contact_col2 = st.columns([1, 1])
            
            with contact_col1:
                if result.get('project_manager'):
                    st.markdown(f"**👤 Gestionnaire:** {result.get('project_manager')}")
                
                # Handle different contact field formats
                contact_email = result.get('contact_email') or result.get('contact')
                if contact_email and contact_email != 'Non spécifié':
                    st.markdown(f"**📧 Email:** {contact_email}")
                
                if result.get('contact_tel') and result.get('contact_tel') != 'Non spécifié':
                    st.markdown(f"**📞 Téléphone:** {result.get('contact_tel')}")
            
            with contact_col2:
                if result.get('contact_info'):
                    st.markdown(f"**📞 Contact:** {result.get('contact_info')}")
                if result.get('contact_adresse') and result.get('contact_adresse') != 'Non spécifié':
                    st.markdown(f"**📍 Adresse:** {result.get('contact_adresse')}")
                if result.get('date_ouverture') and result.get('date_ouverture') != 'Non spécifié':
                    st.markdown(f"**📅 Ouverture:** {result.get('date_ouverture')}")
                if result.get('date_cloture') and result.get('date_cloture') != result.get('deadline'):
                    st.markdown(f"**⏰ Clôture exacte:** {result.get('date_cloture')}")
        
        # Visual Arts specific information
        if (result.get('documents_requis') or result.get('eligibilite') or 
            result.get('duree_projet') or result.get('modalites_versement')):
            st.markdown("---")
            st.markdown("### 📋 **Détails du projet artistique**")
            
            detail_col1, detail_col2 = st.columns([1, 1])
            
            with detail_col1:
                if result.get('duree_projet') and result.get('duree_projet') != 'Non spécifié':
                    st.markdown(f"**⏱️ Durée:** {result.get('duree_projet')}")
                if result.get('modalites_versement') and result.get('modalites_versement') != 'Non spécifié':
                    st.markdown(f"**💳 Modalités:** {result.get('modalites_versement')}")
                if result.get('jury') and result.get('jury') != 'Non spécifié':
                    st.markdown(f"**👥 Jury:** {result.get('jury')}")
                if result.get('selection') and result.get('selection') != 'Non spécifié':
                    st.markdown(f"**🎯 Sélection:** {result.get('selection')}")
            
            with detail_col2:
                if result.get('documents_requis') and result.get('documents_requis'):
                    st.markdown("**📄 Documents requis:**")
                    for doc in result.get('documents_requis'):
                        st.markdown(f"• {doc}")
                
                if result.get('eligibilite') and result.get('eligibilite'):
                    st.markdown("**✅ Critères d'éligibilité:**")
                    for critere in result.get('eligibilite'):
                        st.markdown(f"• {critere}")
                
                if result.get('avantages') and result.get('avantages'):
                    st.markdown("**🎁 Avantages:**")
                    for avantage in result.get('avantages'):
                        st.markdown(f"• {avantage}")
        
        # Action buttons
        st.markdown("---")
        button_col1, button_col2, button_col3 = st.columns([1, 1, 1])
        
        with button_col1:
            # Handle different URL field names
            website_url = result.get('url') or result.get('website')
            if website_url:
                # Display link with validation status
                link_text = "🔗 Ver detalles completos" if result_type == "colombian" else "🔗 Voir les détails complets"
                link_status = result.get('link_status', '')
                
                if result.get('link_active'):
                    st.markdown(f"**[{link_text}]({website_url})**")
                    if link_status:
                        st.markdown(f"<small style='color: green;'>{link_status}</small>", unsafe_allow_html=True)
                else:
                    st.markdown(f"**[{link_text}]({website_url})**")
                    if link_status:
                        st.markdown(f"<small style='color: orange;'>{link_status}</small>", unsafe_allow_html=True)
                    st.markdown("<small style='color: gray;'>⚠️ Link validation recommended</small>", unsafe_allow_html=True)
        
        with button_col2:
            # Handle different contact field names
            contact_email = result.get('contact_email') or result.get('contact')
            if contact_email and '@' in str(contact_email):
                subject_text = "Solicitud de información" if result_type == "colombian" else "Demande d'information"
                mailto_link = f"mailto:{contact_email}?subject={subject_text} - {result.get('title', '')}"
                email_text = "📧 Contactar por email" if result_type == "colombian" else "📧 Contacter par email"
                st.markdown(f"**[{email_text}]({mailto_link})**")
        
        with button_col3:
            # Add to favorites functionality (could be implemented later)
            fav_text = "⭐ Marcar favorito" if result_type == "colombian" else "⭐ Marquer favori"
            if st.button(fav_text, key=f"fav_{index}_{result_type}"):
                success_text = "¡Agregado a favoritos! (funcionalidad próximamente)" if result_type == "colombian" else "Ajouté aux favoris! (fonctionnalité à venir)"
                st.success(success_text)

def main():
    """Main application function"""
    init_session_state()
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>🔍 Subvention Research Bot</h1>
        <p>Discover French & European Financial Aid + Visual Arts Funding</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar for search controls
    with st.sidebar:
        st.markdown("## 🎯 Search Parameters")
        
        # Search type selection
        search_type = st.radio(
            "🎨 Search Type",
            options=["European Subventions", "French Visual Arts", "European Visual Arts", "Colombian Funding", "Colombian Visual Arts"],
            help="Choose between European innovation funding, French visual arts, European visual arts funding, Colombian general funding, or Colombian visual arts funding"
        )
        
        # Keywords input (conditional based on search type)
        keywords = ""  # Initialize default value
        art_type = "All"  # Initialize default value
        country_filter = "all"  # Initialize default value
        min_amount = 0  # Initialize default value
        max_amount = 1000000  # Initialize default value
        
        # Conditional options based on search type
        if search_type == "European Subventions":
            # Keywords input for European subventions
            keywords = st.text_input(
                "🔍 Keywords",
                value="",
                placeholder="e.g., innovation, startup, green energy",
                help="Enter keywords separated by commas for better results"
            )
            
            # Region selection for French subventions
            region = st.selectbox(
                "🇫🇷 French Region",
                options=list(FRENCH_REGIONS.keys()),
                format_func=lambda x: FRENCH_REGIONS[x],
                help="Select a specific French region or leave as 'All regions'"
            )
            
            # Include European results
            include_european = st.checkbox(
                "🇪🇺 Include European Subventions",
                value=True,
                help="Search for European grants and tenders"
            )
            
            # European region selection (only show if European results are included)
            european_region = ""
            if include_european:
                european_region = st.selectbox(
                    "🇪🇺 European Region/Country",
                    options=list(EUROPEAN_REGIONS.keys()),
                    format_func=lambda x: EUROPEAN_REGIONS[x]['name'],
                    help="Select a specific European region/country or leave as 'All European Countries'"
                )
            
            # Colombian options disabled for European Subventions
            include_colombian = False
            colombian_region = ""
            
            # Results limit
            results_limit = st.slider(
                "📊 Maximum Results",
                min_value=10,
                max_value=200,
                value=50,
                step=10,
                help="Number of results to retrieve"
            )
            
        elif search_type == "Colombian Funding":
            # Keywords input for Colombian funding
            keywords = st.text_input(
                "🔍 Keywords",
                value="",
                placeholder="e.g., arte, cultura, emprendimiento, innovación",
                help="Enter keywords in Spanish or English for better results"
            )
            
            # Colombian region selection
            colombian_region = st.selectbox(
                "🇨🇴 Colombian Region/City",
                options=list(COLOMBIAN_REGIONS.keys()),
                format_func=lambda x: COLOMBIAN_REGIONS[x]['name'],
                help="Select a specific Colombian region/city or leave as 'All Colombian Cities'"
            )
            
            # European and French options disabled for Colombian Funding
            region = ""
            include_european = False
            european_region = ""
            include_colombian = True
            
            # Results limit
            results_limit = st.slider(
                "📊 Maximum Results",
                min_value=10,
                max_value=200,
                value=30,
                step=10,
                help="Number of results to retrieve"
            )
            
        elif search_type == "French Visual Arts":
            # Art type selection for French Visual Arts
            art_type = st.selectbox(
                "🎨 Art Type",
                options=["All", "Photographie", "Peinture", "Sculpture", "Arts graphiques", "Illustration", "Céramique"],
                help="Filter by specific art type or see all visual arts opportunities"
            )
            
            # French Visual Arts specific options
            keywords = ""  # Not used for visual arts
            region = "all"
            include_european = False
            european_region = ""
            include_colombian = False
            colombian_region = ""
            results_limit = 50
            
        elif search_type == "European Visual Arts":
            # Art type selection for European Visual Arts
            art_type = st.selectbox(
                "🎨 Art Type",
                options=["All", "Photography", "Visual Arts", "Sculpture", "Installation", "Contemporary Art"],
                help="Filter by specific art type across Europe"
            )
            
            # Country filter for European Visual Arts
            country_filter = st.selectbox(
                "🌍 European Country",
                options=["All Countries", "Germany", "UK", "France", "Italy", "Spain", "Netherlands", 
                        "Austria", "Belgium", "Sweden", "Denmark", "Norway", "Finland", 
                        "Switzerland", "Poland", "Czech Republic", "Portugal"],
                help="Filter by specific European country or see all"
            )
            
            col1, col2 = st.columns(2)
            with col1:
                min_amount = st.number_input(
                    "💰 Min Amount (€)",
                    min_value=0,
                    max_value=1000000,
                    value=0,
                    step=1000,
                    help="Minimum funding amount"
                )
            with col2:
                max_amount = st.number_input(
                    "💰 Max Amount (€)",
                    min_value=0,
                    max_value=1000000,
                    value=1000000,
                    step=1000,
                    help="Maximum funding amount"
                )
            
            # European Visual Arts specific options
            keywords = ""  # Not used for visual arts
            region = "all"
            include_european = False
            european_region = ""
            include_colombian = False
            colombian_region = ""
            results_limit = 50
            
        else:  # Colombian Visual Arts
            # Art type selection for Colombian Visual Arts
            art_type = st.selectbox(
                "🎨 Tipo de Arte",
                options=["All", "Pintura", "Escultura", "Fotografía", "Arte Digital", "Arte Urbano", 
                        "Performance", "Instalación", "Arte Contemporáneo", "Arte Tradicional", "Arte Ambiental"],
                help="Filtrar por tipo específico de arte visual colombiano"
            )
            
            # Colombian city filter
            country_filter = st.selectbox(
                "🇨🇴 Ciudad Colombiana",
                options=["All Cities", "Bogotá", "Medellín", "Cali", "Barranquilla", "Cartagena", 
                        "Bucaramanga", "Pereira", "Armenia", "Manizales", "Eje Cafetero"],
                help="Filtrar por ciudad específica o ver todas las oportunidades"
            )
            
            col1, col2 = st.columns(2)
            with col1:
                min_amount = st.number_input(
                    "💰 Monto Mínimo (COP)",
                    min_value=0,
                    max_value=500000000,
                    value=0,
                    step=5000000,
                    help="Monto mínimo de financiación en pesos colombianos"
                )
            with col2:
                max_amount = st.number_input(
                    "💰 Monto Máximo (COP)",
                    min_value=0,
                    max_value=500000000,
                    value=500000000,
                    step=5000000,
                    help="Monto máximo de financiación en pesos colombianos"
                )
            
            # Colombian Visual Arts specific options
            keywords = ""  # Not used for visual arts
            region = "all"
            include_european = False
            european_region = ""
            include_colombian = False
            colombian_region = ""
            results_limit = 50
        
        # Auto-save options
        auto_save = st.checkbox(
            "💾 Auto-save to Excel",
            value=True,
            help="Automatically save results to Excel file in exports/excel folder"
        )
        
        auto_save_logs = st.checkbox(
            "📊 Auto-save Analysis Logs",
            value=True,
            help="Automatically save detailed analysis and statistics to exports/logs folder"
        )
        
        auto_fix_urls = st.checkbox(
            "🔧 Auto-fix Broken URLs",
            value=False,
            help="Attempt to automatically repair broken URLs using multiple strategies"
        )
        
        st.markdown("---")
        
        # Search button (conditional text based on search type)
        if search_type == "European Subventions":
            search_button_text = "🚀 Search Subventions"
        elif search_type == "French Visual Arts":
            search_button_text = "🎨 Search French Visual Arts"
        elif search_type == "Colombian Funding":
            search_button_text = "🇨🇴 Search Colombian Funding"
        elif search_type == "European Visual Arts":
            search_button_text = "� Search European Visual Arts"
        else:  # Colombian Visual Arts
            search_button_text = "🎨 Buscar Arte Visual Colombiano"
            
        if st.button(search_button_text, type="primary"):
            if search_type == "European Subventions":
                perform_search(search_type, keywords, region, european_region, colombian_region, include_european, include_colombian, results_limit, auto_save, auto_save_logs, auto_fix_urls)
            elif search_type == "Colombian Funding":
                perform_search(search_type, keywords, region, european_region, colombian_region, include_european, include_colombian, results_limit, auto_save, auto_save_logs, auto_fix_urls)
            elif search_type == "French Visual Arts":
                perform_visual_arts_search(art_type, auto_save, auto_save_logs, auto_fix_urls)
            elif search_type == "European Visual Arts":
                perform_european_visual_arts_search(art_type, country_filter, min_amount, max_amount, auto_save, auto_save_logs, auto_fix_urls)
            else:  # Colombian Visual Arts
                perform_colombian_visual_arts_search(art_type, country_filter, min_amount, max_amount, auto_save, auto_save_logs, auto_fix_urls)
        
        # Manual save button (if auto-save is disabled)
        if not auto_save and st.session_state.search_performed:
            st.markdown("### 💾 Export Options")
            if st.button("📥 Save to Excel"):
                save_results_to_excel()
        
        # Clear results button
        if st.session_state.search_performed:
            if st.button("🗑️ Clear Results"):
                st.session_state.search_results = {'french': [], 'european': []}
                st.session_state.search_performed = False
                st.rerun()
    
    # Main content area
    if st.session_state.search_performed:
        display_search_results(st.session_state.search_results)
        
        # Download section
        st.markdown("---")
        st.markdown("### 📥 Download Results")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("📊 Download Full Excel Report", help="Download complete results including all projects"):
                save_results_to_excel(download=True)
        
        with col2:
            if st.button("✅ Download Working URLs Only", help="Download only projects with verified working URLs"):
                save_filtered_excel(filter_type="working_urls")
        
        with col3:
            if st.button("📋 Generate Analysis Report"):
                save_results_to_excel_and_logs(False, True, False)
        
        # File organization info
        st.markdown("---")
        st.markdown("### 📁 **File Organization**")
        st.info("""
        **Your files are automatically organized in:**
        - 📊 **Excel Reports**: `exports/excel/` - Detailed spreadsheets with all funding data
        - ✅ **Filtered Excel**: `exports/excel/` - Curated reports with only working URLs
        - 📋 **Analysis Logs**: `exports/logs/` - JSON and text analysis with statistics and recommendations
        - � **URL Fix Logs**: `exports/fixed_logs/` - Detailed URL repair attempts and results
        """)
        
        # Add download tips
        st.markdown("### 💡 **Download Tips**")
        st.success("""
        **📥 Download Options Available:**
        - **Full Excel Report**: Complete dataset with all projects
        - **Working URLs Only**: Filtered dataset with verified functional links
        - **Analysis Report**: Comprehensive statistics and URL validation insights
        
        **🔍 URL Quality Features:**
        - Automatic URL validation during analysis
        - Broken URL detection and categorization
        - Optional URL repair attempts with detailed logs
        """)
        
        # Show recent files if any exist
        try:
            excel_files = []
            log_files = []
            
            if os.path.exists("exports/excel"):
                excel_files = [f for f in os.listdir("exports/excel") if f.endswith('.xlsx')][-3:]  # Last 3 files
            if os.path.exists("exports/logs"):
                log_files = [f for f in os.listdir("exports/logs") if f.endswith('.json')][-3:]  # Last 3 files
            
            if excel_files or log_files:
                st.markdown("#### 📋 **Recent Files**")
                
                if excel_files:
                    st.markdown("**📊 Recent Excel Reports:**")
                    for file in reversed(excel_files):  # Show newest first
                        st.markdown(f"  • `exports/excel/{file}`")
                
                if log_files:
                    st.markdown("**📋 Recent Analysis Logs:**")
                    for file in reversed(log_files):  # Show newest first
                        st.markdown(f"  • `exports/logs/{file}`")
                        
        except Exception as e:
            pass  # Ignore file listing errors
    
    else:
        # Welcome message
        st.markdown("""
        <div class="welcome-container">
            <h3>Welcome to the Enhanced Subvention Research Bot! 🎉</h3>
            <p>This comprehensive tool helps you discover financial aid and grant opportunities from:</p>
            <ul>
                <li><strong>🇫🇷 French Sources:</strong> Aides-Territoires API with deep regional filtering</li>
                <li><strong>🇪🇺 European Sources:</strong> TED Europa and EU Funding Portal for innovation</li>
                <li><strong>🎨 French Visual Arts:</strong> Specialized French funding for photography, painting, sculpture</li>
                <li><strong>🌍 European Visual Arts:</strong> Comprehensive EU-wide funding with real-time API data, amounts up to €250,000+</li>
            </ul>
            <p>Use the sidebar to choose your search type and configure parameters, then click the search button to get started!</p>
        </div>
        """, unsafe_allow_html=True)
        
        # API Status Information
        st.info("""
        📡 **API Status**: Some external APIs may require authentication or have rate limits. 
        When live data is unavailable, the bot provides curated sample results to demonstrate functionality.
        """)
        
        # Sample searches
        st.markdown("### 💡 Sample Searches")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🚀 Innovation & Startups"):
                perform_search("European Subventions", "innovation, startup, entrepreneur", "fr-idf", "", "", True, False, 30, True, True, False)
        
        with col2:
            if st.button("🌱 Green Energy"):
                perform_search("European Subventions", "renewable energy, green, environment", "", "", "", True, False, 30, True, True, False)
        
        with col3:
            if st.button("🎓 Education & Research"):
                perform_search("European Subventions", "education, research, university", "", "", "", True, False, 30, True, True, False)

def perform_search(search_type: str, keywords: str, region: str, european_region: str, colombian_region: str, include_european: bool, include_colombian: bool, limit: int, auto_save: bool, auto_save_logs: bool, auto_fix_urls: bool = False):
    """Perform the subvention search"""
    try:
        # Show progress
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("🔍 Initializing search...")
        progress_bar.progress(10)
        
        # Initialize searcher
        searcher = SubventionSearcher()
        
        # Perform search based on search type
        if search_type == "Colombian Funding":
            status_text.text("�🇴 Searching Colombian funding opportunities...")
            progress_bar.progress(50)
            
            # Only search Colombian sources
            colombian_results = searcher.search_colombian_funding(
                keywords=keywords,
                region=colombian_region,
                limit=limit
            )
            results = {'french': [], 'european': [], 'colombian': colombian_results}
        else:
            # Regular European/French search
            status_text.text("�🇫🇷 Searching French subventions...")
            progress_bar.progress(30)
            
            # Perform comprehensive search
            results = searcher.search_all(
                keywords=keywords,
                region=region,
                european_region=european_region,
                colombian_region=colombian_region,
                include_european=include_european,
                include_colombian=include_colombian,
                limit=limit
            )
        
        progress_bar.progress(80)
        status_text.text("📊 Processing results...")
        
        # Store results in session state
        st.session_state.search_results = results
        st.session_state.last_search_params = {
            'search_type': search_type,
            'keywords': keywords,
            'region': region,
            'european_region': european_region,
            'colombian_region': colombian_region,
            'include_european': include_european,
            'include_colombian': include_colombian,
            'limit': limit
        }
        st.session_state.search_performed = True
        
        progress_bar.progress(90)
        
        # Auto-save if enabled
        if auto_save or auto_save_logs:
            status_text.text("💾 Saving results and analysis...")
            save_results_to_excel_and_logs(auto_save, auto_save_logs, auto_fix_urls)
        
        progress_bar.progress(100)
        status_text.text("✅ Search completed!")
        
        # Clear progress indicators
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        # Show success message
        total_results = len(results['french']) + len(results['european'])
        st.success(f"🎉 Found {total_results} subventions! ({len(results['french'])} French, {len(results['european'])} European)")
        
        # Rerun to show results
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Search failed: {str(e)}")
        logger.error(f"Search error: {e}")

def perform_visual_arts_search(art_type: str, auto_save: bool, auto_save_logs: bool, auto_fix_urls: bool = False):
    """Perform the visual arts funding search"""
    try:
        # Show progress
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("🎨 Initializing visual arts search...")
        progress_bar.progress(10)
        
        # Initialize visual arts searcher
        arts_searcher = RechercheArtsVisuels()
        
        status_text.text("🎨 Searching visual arts funding opportunities...")
        progress_bar.progress(50)
        
        # Perform search with filtering
        if art_type == "All":
            all_projects = arts_searcher.rechercher_arts_visuels()
        else:
            all_projects = arts_searcher.rechercher_arts_visuels(type_art=art_type)
        
        progress_bar.progress(80)
        status_text.text("📊 Processing visual arts results...")
        
        # Convert to the format expected by the display function
        formatted_results = []
        for project in all_projects:
            formatted_project = {
                'title': project['titre'],
                'description': project['description'],
                'organisme': project['organisme'],
                'montant': project['montant'],
                'deadline': project['deadline'],
                'url': project['url'],
                'type': project['type_projet'],
                'statut': project['statut'],
                'criteres': ', '.join(project['criteres']),
                'source': 'French Visual Arts',
                # Additional detailed information
                'contact_email': project.get('contact_email', 'Non spécifié'),
                'contact_tel': project.get('contact_tel', 'Non spécifié'),
                'contact_adresse': project.get('contact_adresse', 'Non spécifié'),
                'date_ouverture': project.get('date_ouverture', 'Non spécifié'),
                'date_cloture': project.get('date_cloture', project['deadline']),
                'documents_requis': project.get('documents_requis', []),
                'eligibilite': project.get('eligibilite', []),
                'duree_projet': project.get('duree_projet', 'Non spécifié'),
                'modalites_versement': project.get('modalites_versement', 'Non spécifié'),
                'jury': project.get('jury', 'Non spécifié'),
                'avantages': project.get('avantages', []),
                'selection': project.get('selection', 'Non spécifié')
            }
            formatted_results.append(formatted_project)
        
        # Store results in session state (format compatible with existing display)
        visual_arts_results = {
            'french': formatted_results,
            'european': []  # No European results for visual arts
        }
        
        st.session_state.search_results = visual_arts_results
        st.session_state.last_search_params = {
            'search_type': 'French Visual Arts',
            'art_type': art_type,
            'keywords': f"Visual Arts - {art_type}",
            'region': 'France',
            'european_region': '',
            'include_european': False,
            'limit': len(formatted_results)
        }
        st.session_state.search_performed = True
        
        progress_bar.progress(90)
        
        # Auto-save if enabled
        if auto_save or auto_save_logs:
            status_text.text("💾 Saving visual arts results...")
            save_results_to_excel_and_logs(auto_save, auto_save_logs, auto_fix_urls)
        
        progress_bar.progress(100)
        status_text.text("✅ Visual arts search completed!")
        
        # Clear progress indicators
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        # Show success message
        total_results = len(formatted_results)
        st.success(f"🎨 Found {total_results} visual arts funding opportunities!")
        
        # Rerun to show results
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Visual arts search failed: {str(e)}")
        logger.error(f"Visual arts search error: {e}")

def perform_european_visual_arts_search(art_type: str, country_filter: str, min_amount: int, max_amount: int, auto_save: bool, auto_save_logs: bool, auto_fix_urls: bool = False):
    """Perform the comprehensive European visual arts funding search"""
    try:
        # Show progress
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("🌍 Initializing European visual arts search...")
        progress_bar.progress(10)
        
        # Initialize European arts searcher
        eu_arts_searcher = EuropeanVisualArtsFunding()
        
        status_text.text("🎨 Searching European visual arts funding...")
        progress_bar.progress(30)
        
        # Prepare filters
        country = None if country_filter == "All Countries" else country_filter.lower().replace(" ", "_")
        art_filter = None if art_type == "All" else art_type
        min_amt = min_amount if min_amount > 0 else None
        max_amt = max_amount if max_amount < 1000000 else None
        
        # Perform search
        all_projects = eu_arts_searcher.search_european_visual_arts_funding(
            country=country,
            art_type=art_filter,
            min_amount=min_amt,
            max_amount=max_amt
        )
        
        progress_bar.progress(70)
        status_text.text("📊 Processing European results...")
        
        # Convert to the format expected by the display function
        formatted_results = []
        for project in all_projects:
            # Calculate amount in EUR for consistency
            amount_min = project.get('montant_min', 0)
            amount_max = project.get('montant_max', 0)
            currency = project.get('currency', 'EUR')
            
            # Convert to EUR if needed (simplified conversion)
            if currency == 'GBP':
                amount_min = int(amount_min * 1.17)  # Approximate EUR conversion
                amount_max = int(amount_max * 1.17)
            elif currency == 'DKK':
                amount_min = int(amount_min * 0.134)
                amount_max = int(amount_max * 0.134)
            
            formatted_project = {
                'title': project['title'],
                'description': project['description'],
                'organisme': project['organisme'],
                'montant': project['montant'],
                'amount_min': str(amount_min),
                'amount_max': str(amount_max),
                'deadline': project['deadline'],
                'url': project['url'],
                'type': project['type_projet'],
                'statut': project['statut'],
                'source': f"European Visual Arts - {project['country']}",
                'perimeter': project['country'],
                # Enhanced European information
                'contact_email': project.get('contact_email', 'Not specified'),
                'contact_tel': project.get('contact_tel', 'Not specified'),
                'contact_adresse': project.get('contact_adresse', 'Not specified'),
                'date_ouverture': project.get('date_ouverture', 'Not specified'),
                'date_cloture': project.get('date_cloture', project['deadline']),
                'documents_requis': project.get('documents_requis', []),
                'eligibilite': project.get('eligibilite', []),
                'duree_projet': project.get('duree_projet', 'Not specified'),
                'modalites_versement': project.get('modalites_versement', 'Not specified'),
                'co_financing': project.get('co_financing', 'Not specified'),
                'additional_support': project.get('additional_support', 'Not specified'),
                'jury': project.get('jury', 'Not specified'),
                'selection_criteria': project.get('selection_criteria', 'Not specified'),
                'currency': currency,
                'original_amount': project['montant']
            }
            formatted_results.append(formatted_project)
        
        # Store results in session state
        european_arts_results = {
            'french': [],  # No French results for European search
            'european': formatted_results
        }
        
        st.session_state.search_results = european_arts_results
        st.session_state.last_search_params = {
            'search_type': 'European Visual Arts',
            'art_type': art_type,
            'country': country_filter,
            'min_amount': min_amount,
            'max_amount': max_amount,
            'keywords': f"European Visual Arts - {art_type} - {country_filter}",
            'region': 'Europe',
            'european_region': country_filter,
            'include_european': True,
            'limit': len(formatted_results)
        }
        st.session_state.search_performed = True
        
        progress_bar.progress(90)
        
        # Auto-save if enabled
        if auto_save or auto_save_logs:
            status_text.text("💾 Saving European visual arts results...")
            save_results_to_excel_and_logs(auto_save, auto_save_logs, auto_fix_urls)
        
        progress_bar.progress(100)
        status_text.text("✅ European visual arts search completed!")
        
        # Clear progress indicators
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        # Show success message with funding statistics
        total_results = len(formatted_results)
        total_funding = sum(int(r.get('amount_max', '0').replace(',', '')) for r in formatted_results if r.get('amount_max', '').replace(',', '').isdigit())
        
        st.success(f"🌍 Found {total_results} European visual arts opportunities! Total funding available: €{total_funding:,}")
        
        # Rerun to show results
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ European visual arts search failed: {str(e)}")
        logger.error(f"European visual arts search error: {e}")

def perform_colombian_visual_arts_search(art_type: str, country_filter: str, min_amount: int, max_amount: int, auto_save: bool, auto_save_logs: bool, auto_fix_urls: bool = False):
    """Perform the comprehensive Colombian visual arts funding search"""
    try:
        # Show progress
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("🇨🇴 Iniciando búsqueda de arte visual colombiano...")
        progress_bar.progress(10)
        
        # Initialize Colombian arts searcher
        colombian_arts_searcher = ColombianVisualArtsFunding()
        
        status_text.text("🎨 Buscando financiación de arte visual colombiano...")
        progress_bar.progress(30)
        
        # Prepare filters
        city = None if country_filter == "All Cities" else country_filter.lower().replace(" ", "_")
        art_filter = None if art_type == "All" else art_type
        min_amt = min_amount if min_amount > 0 else None
        max_amt = max_amount if max_amount < 500000000 else None
        
        # Perform search
        all_projects = colombian_arts_searcher.search_colombian_visual_arts_funding(
            city=city,
            art_type=art_filter,
            min_amount=min_amt,
            max_amount=max_amt
        )
        
        progress_bar.progress(70)
        status_text.text("📊 Procesando resultados colombianos...")
        
        # Convert to the format expected by the display function
        formatted_results = []
        for project in all_projects:
            # Calculate amount in COP for consistency
            amount_min = project.get('montant_min', 0)
            amount_max = project.get('montant_max', 0)
            currency = project.get('currency', 'COP')
            
            # Convert EUR to COP if needed
            if currency == 'EUR':
                amount_min = int(amount_min * 4500)  # Approximate COP conversion
                amount_max = int(amount_max * 4500)
            
            formatted_project = {
                'title': project['title'],
                'description': project['description'],
                'organization': project['organisme'],
                'montant': project['montant'],
                'funding_amount': project['montant'],
                'amount_min': str(amount_min),
                'amount_max': str(amount_max),
                'deadline': project['deadline'],
                'days_until_deadline': project.get('days_until_deadline', 999),
                'url': project.get('url', ''),
                'website': project.get('url', ''),
                'type': project['type_projet'],
                'statut': project['statut'],
                'source': f"Colombian Visual Arts - {project['city']}",
                'perimeter': project['city'],
                'location': project['city'],
                'category': 'Arte Visual Colombiano',
                # Enhanced Colombian information
                'contact_email': project.get('contact_email', 'No especificado'),
                'contact': project.get('contact_email', 'No especificado'),
                'contact_tel': project.get('contact_tel', 'No especificado'),
                'contact_phone': project.get('contact_tel', 'No especificado'),
                'contact_adresse': project.get('contact_adresse', 'No especificado'),
                'date_ouverture': project.get('date_ouverture', 'No especificado'),
                'date_cloture': project.get('date_cloture', project['deadline']),
                'documents_requis': project.get('documents_requis', []),
                'eligibilite': project.get('eligibilite', []),
                'eligibility': ', '.join(project.get('eligibilite', [])) if project.get('eligibilite') else 'Ver convocatoria',
                'duree_projet': project.get('duree_projet', 'No especificado'),
                'modalites_versement': project.get('modalites_versement', 'No especificado'),
                'co_financing': project.get('co_financing', 'No especificado'),
                'additional_support': project.get('additional_support', 'No especificado'),
                'art_focus': project.get('art_focus', []),
                'currency': currency,
                'original_amount': project['montant'],
                'urgency': project.get('urgency', 'medium')
            }
            formatted_results.append(formatted_project)
        
        # Store results in session state
        colombian_arts_results = {
            'french': [],  # No French results for Colombian search
            'european': [],  # No European results for Colombian search
            'colombian': formatted_results
        }
        
        st.session_state.search_results = colombian_arts_results
        st.session_state.last_search_params = {
            'search_type': 'Colombian Visual Arts',
            'art_type': art_type,
            'city': country_filter,
            'min_amount': min_amount,
            'max_amount': max_amount,
            'keywords': f"Colombian Visual Arts - {art_type} - {country_filter}",
            'region': 'Colombia',
            'european_region': '',
            'include_european': False,
            'include_colombian': True,
            'limit': len(formatted_results)
        }
        st.session_state.search_performed = True
        
        progress_bar.progress(90)
        
        # Auto-save if enabled
        if auto_save or auto_save_logs:
            status_text.text("💾 Guardando resultados de arte visual colombiano...")
            save_results_to_excel_and_logs(auto_save, auto_save_logs, auto_fix_urls)
        
        progress_bar.progress(100)
        status_text.text("✅ ¡Búsqueda de arte visual colombiano completada!")
        
        # Clear progress indicators
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        # Show success message with funding statistics
        total_results = len(formatted_results)
        total_funding = sum(int(r.get('amount_max', '0').replace(',', '')) for r in formatted_results if r.get('amount_max', '').replace(',', '').isdigit())
        
        st.success(f"🇨🇴 ¡Encontradas {total_results} oportunidades de arte visual colombiano! Financiación total disponible: COP ${total_funding:,}")
        
        # Rerun to show results
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Colombian visual arts search failed: {str(e)}")
        logger.error(f"Colombian visual arts search error: {e}")

def save_results_to_excel_and_logs(auto_save: bool, auto_save_logs: bool, auto_fix_urls: bool = False):
    """Save search results to Excel and/or create analysis logs"""
    try:
        results = st.session_state.search_results
        search_params = st.session_state.get('last_search_params', {})
        all_results = results.get('french', []) + results.get('european', []) + results.get('colombian', [])
        
        if not all_results:
            st.warning("⚠️ No results to save!")
            return
        
        # Get search keywords for filename
        keywords = search_params.get('keywords', '')
        search_type = search_params.get('search_type', 'search')
        
        # Clean keywords for filename (remove special characters, limit length)
        if keywords:
            clean_keywords = re.sub(r'[^\w\s-]', '', keywords)
            clean_keywords = re.sub(r'\s+', '_', clean_keywords.strip())
            clean_keywords = clean_keywords[:30] if len(clean_keywords) > 30 else clean_keywords
            keyword_part = f"{clean_keywords}_"
        else:
            keyword_part = f"{search_type.lower().replace(' ', '_')}_"
        
        # Create timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        saved_files = []
        
        # Save to Excel if requested
        if auto_save:
            excel_filename = f"subventions_{keyword_part}{timestamp}.xlsx"
            df = pd.DataFrame(all_results)
            excel_file = create_styled_excel(df, excel_filename)
            if excel_file:
                saved_files.append(f"📊 Excel: {excel_file}")
        
        # Save analysis logs if requested
        if auto_save_logs:
            log_filename = f"analysis_{keyword_part}{timestamp}.json"
            log_file = create_analysis_log(results, search_params, log_filename, auto_fix_urls)
            if log_file:
                saved_files.append(f"📋 Analysis: {log_file}")
                # Also mention the text version
                text_log = log_file.replace('.json', '.txt')
                saved_files.append(f"📝 Summary: {text_log}")
        
        # Show success message with file locations
        if saved_files:
            st.success("✅ Files saved successfully!")
            for file_info in saved_files:
                st.info(file_info)
            
            # Show analysis summary if logs were created
            if auto_save_logs and all_results:
                total_results = len(all_results)
                urgent_count = len([r for r in all_results if r.get('days_until_deadline', 999) <= 30])
                total_funding = sum([float(r.get('amount_max', '0').replace(',', '')) for r in all_results if r.get('amount_max', '').replace(',', '').isdigit()])
                
                st.info(f"📊 Analysis Summary: {total_results} opportunities, {urgent_count} urgent, €{total_funding:,.0f} total funding")
        
    except Exception as e:
        st.error(f"❌ Failed to save files: {str(e)}")
        logger.error(f"Save error: {e}")

def save_results_to_excel(download: bool = False):
    """Save search results to Excel file"""
    try:
        results = st.session_state.search_results
        all_results = results.get('french', []) + results.get('european', []) + results.get('colombian', [])
        
        if not all_results:
            st.warning("⚠️ No results to save!")
            return
        
        # Convert to DataFrame
        df = pd.DataFrame(all_results)
        
        # Get search keywords for filename
        search_params = st.session_state.get('last_search_params', {})
        keywords = search_params.get('keywords', '')
        
        # Clean keywords for filename (remove special characters, limit length)
        if keywords:
            # Remove special characters and replace spaces with underscores
            clean_keywords = re.sub(r'[^\w\s-]', '', keywords)
            clean_keywords = re.sub(r'\s+', '_', clean_keywords.strip())
            # Limit length to avoid very long filenames
            clean_keywords = clean_keywords[:30] if len(clean_keywords) > 30 else clean_keywords
            keyword_part = f"{clean_keywords}_"
        else:
            keyword_part = ""
        
        # Create filename with keywords and timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"subventions_{keyword_part}{timestamp}.xlsx"
        
        # Create styled Excel file in organized directory
        excel_file = create_styled_excel(df, filename)
        
        if excel_file:
            if download:
                # Provide download link
                with open(excel_file, 'rb') as file:
                    st.download_button(
                        label="📥 Download Excel File",
                        data=file.read(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            st.success(f"✅ Results saved to: {excel_file}")
            st.info(f"📊 Saved {len(all_results)} subventions to Excel file")
        else:
            st.error("❌ Failed to create Excel file")
            
    except Exception as e:
        st.error(f"❌ Failed to save Excel file: {str(e)}")
        logger.error(f"Excel save error: {e}")

# Footer
def show_footer():
    """Display application footer"""
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; padding: 1rem;">
        <p>🔍 <strong>Subvention Research Bot</strong> | Made with ❤️ using Streamlit</p>
        <p>Sources: Aides-Territoires API • TED Europa • EU Funding Portal • European Visual Arts Network</p>
        <p>📁 Files organized in: <code>exports/excel/</code> • <code>exports/logs/</code> • <code>exports/analysis/</code></p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
    show_footer()
